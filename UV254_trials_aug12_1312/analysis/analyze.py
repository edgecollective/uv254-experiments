"""Analyze Aug 12 1312 dataset: Ver 7.0 (2-pt affine) validation runs.

Two blocks of interest:
  Block A (cols J-N, rows 55-62): Ver 7.0 firmware on fluorescein solutions,
    calibrated with 0.010 and 0.100 fluorescein standards.
  Block B (cols R-X, rows 55-83): head-to-head V6.1 vs V7.0. Cols R,S,T =
    RealTech, DIY raw, DIY corr for V6.1. Cols V,W,X = same for V7.0.

Output plots + summary.md/csv in this folder.
"""
import csv
from pathlib import Path
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

HERE = Path(__file__).parent
XLSX = HERE.parent / "data.xlsx"

wb = load_workbook(XLSX, data_only=True)
ws = wb["Sheet1"]

def col_letter_to_idx(letter):
    return ord(letter.upper()) - ord('A') + 1

def read_block(cols, rows):
    """Read rows for given columns, skipping any row where all cells are None."""
    out = []
    for r in rows:
        vals = []
        for c in cols:
            cell = ws.cell(row=r, column=col_letter_to_idx(c))
            vals.append(cell.value)
        if all(v is None for v in vals):
            continue
        out.append(vals)
    return out

# Block A: Ver 7.0 fluorescein-only
# cols J=RT, K=trans, L=V, M=raw abs, N=corr abs; rows 55-62
block_a = read_block(['J', 'M', 'N'], range(55, 63))
block_a = [(rt, raw, corr) for rt, raw, corr in block_a
           if rt is not None and raw is not None and corr is not None]
print("Block A (Ver 7.0 fluor-only):", block_a)

# Block B v6.1: cols R=RT, S=raw, T=corr; rows 55-70
block_b_v61 = read_block(['R', 'S', 'T'], range(55, 71))
block_b_v61 = [(rt, raw, corr) for rt, raw, corr in block_b_v61
               if rt is not None and raw is not None and corr is not None]
print("Block B V6.1:", block_b_v61)

# Block B v7.0: cols V=RT, W=raw, X=corr; rows 55-83
block_b_v70 = read_block(['V', 'W', 'X'], range(55, 84))
block_b_v70 = [(rt, raw, corr) for rt, raw, corr in block_b_v70
               if rt is not None and raw is not None and corr is not None]
print("Block B V7.0:", block_b_v70)

def fit_line(x, y):
    """Return slope, intercept, r2 for y = m*x + b."""
    x = np.array(x); y = np.array(y)
    if len(x) < 2: return None, None, None
    m, b = np.polyfit(x, y, 1)
    yhat = m*x + b
    ss_res = np.sum((y - yhat)**2)
    ss_tot = np.sum((y - np.mean(y))**2)
    r2 = 1 - ss_res/ss_tot if ss_tot > 0 else None
    return m, b, r2

def residual_stats(rt, corr):
    """residuals = corr - rt; return (MAE, max_abs, mean_signed, all-positive?)."""
    r = np.array(corr) - np.array(rt)
    return (
        float(np.mean(np.abs(r))),
        float(np.max(np.abs(r))),
        float(np.mean(r)),
        bool(np.all(r > 0)),
        r.tolist(),
    )

def back_out_affine(raws, corrs):
    """From raw and corr pairs, recover m,b (should be exactly linear)."""
    raws = np.array(raws); corrs = np.array(corrs)
    # Use least-squares fit; will be near-perfect if firmware applied affine.
    m, b = np.polyfit(raws, corrs, 1)
    return float(m), float(b)

# ----- Analyze Block A -----
rt_a  = [x[0] for x in block_a]
raw_a = [x[1] for x in block_a]
corr_a = [x[2] for x in block_a]

m_a_raw, b_a_raw, r2_a_raw = fit_line(rt_a, raw_a)
m_a_corr, b_a_corr, r2_a_corr = fit_line(rt_a, corr_a)
mae_a_raw, max_a_raw, mean_a_raw, allpos_a_raw, _ = residual_stats(rt_a, raw_a)
mae_a_corr, max_a_corr, mean_a_corr, allpos_a_corr, _ = residual_stats(rt_a, corr_a)
m_fw_a, b_fw_a = back_out_affine(raw_a, corr_a)

print(f"\nBlock A (Ver 7.0 fluor-only):")
print(f"  RAW  slope={m_a_raw:.3f}, intercept={b_a_raw:.5f}, r2={r2_a_raw:.4f}")
print(f"  CORR slope={m_a_corr:.3f}, intercept={b_a_corr:.5f}, r2={r2_a_corr:.4f}")
print(f"  RAW  MAE={mae_a_raw*1000:.1f} mAU, max={max_a_raw*1000:.1f} mAU, mean_signed={mean_a_raw*1000:+.1f}")
print(f"  CORR MAE={mae_a_corr*1000:.1f} mAU, max={max_a_corr*1000:.1f} mAU, mean_signed={mean_a_corr*1000:+.1f}, all_pos={allpos_a_corr}")
print(f"  Firmware affine params: m={m_fw_a:.4f}, b={b_fw_a:.5f} ({b_fw_a*1000:+.2f} mAU)")

# ----- Analyze Block B v6.1 -----
rt_b61  = [x[0] for x in block_b_v61]
raw_b61 = [x[1] for x in block_b_v61]
corr_b61 = [x[2] for x in block_b_v61]

m_b61_raw, b_b61_raw, r2_b61_raw = fit_line(rt_b61, raw_b61)
m_b61_corr, b_b61_corr, r2_b61_corr = fit_line(rt_b61, corr_b61)
mae_b61_raw, max_b61_raw, mean_b61_raw, _, res_b61_raw = residual_stats(rt_b61, raw_b61)
mae_b61_corr, max_b61_corr, mean_b61_corr, allpos_b61, res_b61 = residual_stats(rt_b61, corr_b61)
k_b61_avg = float(np.mean([c/r if r > 0 else np.nan for c, r in zip(corr_b61, raw_b61)]))

print(f"\nBlock B (Ver 6.1 head-to-head side):")
print(f"  RAW  slope={m_b61_raw:.3f}, intercept={b_b61_raw:.5f}, r2={r2_b61_raw:.4f}")
print(f"  CORR slope={m_b61_corr:.3f}, intercept={b_b61_corr:.5f}, r2={r2_b61_corr:.4f}")
print(f"  RAW  MAE={mae_b61_raw*1000:.1f} mAU")
print(f"  CORR MAE={mae_b61_corr*1000:.1f} mAU, max={max_b61_corr*1000:.1f} mAU, mean_signed={mean_b61_corr*1000:+.1f}, all_pos={allpos_b61}")
print(f"  k (mean corr/raw): {k_b61_avg:.3f}")

# ----- Analyze Block B v7.0 -----
rt_b70  = [x[0] for x in block_b_v70]
raw_b70 = [x[1] for x in block_b_v70]
corr_b70 = [x[2] for x in block_b_v70]

m_b70_raw, b_b70_raw, r2_b70_raw = fit_line(rt_b70, raw_b70)
m_b70_corr, b_b70_corr, r2_b70_corr = fit_line(rt_b70, corr_b70)
mae_b70_raw, max_b70_raw, mean_b70_raw, _, res_b70_raw = residual_stats(rt_b70, raw_b70)
mae_b70_corr, max_b70_corr, mean_b70_corr, allpos_b70, res_b70 = residual_stats(rt_b70, corr_b70)
m_fw_b70, b_fw_b70 = back_out_affine(raw_b70, corr_b70)

print(f"\nBlock B (Ver 7.0 head-to-head side):")
print(f"  RAW  slope={m_b70_raw:.3f}, intercept={b_b70_raw:.5f}, r2={r2_b70_raw:.4f}")
print(f"  CORR slope={m_b70_corr:.3f}, intercept={b_b70_corr:.5f}, r2={r2_b70_corr:.4f}")
print(f"  RAW  MAE={mae_b70_raw*1000:.1f} mAU")
print(f"  CORR MAE={mae_b70_corr*1000:.1f} mAU, max={max_b70_corr*1000:.1f} mAU, mean_signed={mean_b70_corr*1000:+.1f}, all_pos={allpos_b70}")
print(f"  Firmware affine params: m={m_fw_b70:.4f}, b={b_fw_b70:.5f} ({b_fw_b70*1000:+.2f} mAU)")

# ====================================================
# Plots
# ====================================================

# --- Plot 1: Block A (Ver 7.0 fluor-only) ---
fig, axes = plt.subplots(1, 2, figsize=(12, 5))
ax = axes[0]
xmax = max(max(rt_a), max(raw_a), max(corr_a)) * 1.05
xs = np.linspace(0, xmax, 100)
ax.plot(xs, xs, 'k--', lw=1, alpha=0.5, label='y = x (perfect)')
ax.scatter(rt_a, raw_a, c='tab:blue', s=60, label=f'DIY raw (slope {m_a_raw:.3f})', zorder=3)
ax.scatter(rt_a, corr_a, c='tab:red', s=60, marker='s', label=f'DIY corr (slope {m_a_corr:.3f})', zorder=3)
ax.plot(xs, m_a_raw*xs + b_a_raw, 'tab:blue', alpha=0.5, lw=1)
ax.plot(xs, m_a_corr*xs + b_a_corr, 'tab:red', alpha=0.5, lw=1)
ax.set_xlabel('RealTech absorbance (cm$^{-1}$)')
ax.set_ylabel('DIY absorbance (cm$^{-1}$)')
ax.set_title(f'Ver 7.0 on fluorescein (n={len(rt_a)})\naffine m={m_fw_a:.3f}, b={b_fw_a*1000:+.1f} mAU')
ax.set_xlim(0, xmax); ax.set_ylim(0, xmax)
ax.legend(loc='upper left', fontsize=9)
ax.grid(alpha=0.3); ax.set_aspect('equal')

ax = axes[1]
res_a_raw = np.array(corr_a) * 0 + (np.array(raw_a) - np.array(rt_a))
res_a_corr = np.array(corr_a) - np.array(rt_a)
ax.axhline(0, color='k', lw=1, alpha=0.5)
ax.scatter(rt_a, res_a_raw*1000, c='tab:blue', s=60, label=f'raw residual (MAE {mae_a_raw*1000:.1f} mAU)', zorder=3)
ax.scatter(rt_a, res_a_corr*1000, c='tab:red', s=60, marker='s', label=f'corr residual (MAE {mae_a_corr*1000:.1f} mAU)', zorder=3)
ax.set_xlabel('RealTech absorbance (cm$^{-1}$)')
ax.set_ylabel('DIY - RealTech (mAU)')
ax.set_title('Ver 7.0 fluor: residuals (corr swings both signs — good)')
ax.legend(loc='best', fontsize=9)
ax.grid(alpha=0.3)

plt.tight_layout()
plt.savefig(HERE / 'plot_v70_fluor_only.png', dpi=140, bbox_inches='tight')
plt.close()

# --- Plot 2: Block B head-to-head (V6.1 vs V7.0) ---
fig, axes = plt.subplots(2, 2, figsize=(13, 10))

# Top left: V6.1 scatter
ax = axes[0, 0]
xmax = max(max(rt_b61), max(corr_b61)) * 1.1
xs = np.linspace(0, xmax, 100)
ax.plot(xs, xs, 'k--', lw=1, alpha=0.5, label='y = x')
ax.scatter(rt_b61, raw_b61, c='tab:blue', s=50, label=f'raw (slope {m_b61_raw:.3f})', zorder=3)
ax.scatter(rt_b61, corr_b61, c='tab:red', s=50, marker='s', label=f'corr (slope {m_b61_corr:.3f})', zorder=3)
ax.plot(xs, m_b61_corr*xs + b_b61_corr, 'tab:red', alpha=0.4, lw=1)
ax.set_xlabel('RealTech (cm$^{-1}$)'); ax.set_ylabel('DIY (cm$^{-1}$)')
ax.set_title(f'Ver 6.1 (1-pt cal, k≈{k_b61_avg:.3f})  n={len(rt_b61)}')
ax.set_xlim(0, xmax); ax.set_ylim(0, xmax); ax.set_aspect('equal')
ax.legend(loc='upper left', fontsize=9); ax.grid(alpha=0.3)

# Top right: V7.0 scatter
ax = axes[0, 1]
xmax = max(max(rt_b70), max(corr_b70)) * 1.1
xs = np.linspace(0, xmax, 100)
ax.plot(xs, xs, 'k--', lw=1, alpha=0.5, label='y = x')
ax.scatter(rt_b70, raw_b70, c='tab:blue', s=50, label=f'raw (slope {m_b70_raw:.3f})', zorder=3)
ax.scatter(rt_b70, corr_b70, c='tab:red', s=50, marker='s', label=f'corr (slope {m_b70_corr:.3f})', zorder=3)
ax.plot(xs, m_b70_corr*xs + b_b70_corr, 'tab:red', alpha=0.4, lw=1)
ax.set_xlabel('RealTech (cm$^{-1}$)'); ax.set_ylabel('DIY (cm$^{-1}$)')
ax.set_title(f'Ver 7.0 (2-pt affine, m={m_fw_b70:.3f}, b={b_fw_b70*1000:+.1f} mAU)  n={len(rt_b70)}')
ax.set_xlim(0, xmax); ax.set_ylim(0, xmax); ax.set_aspect('equal')
ax.legend(loc='upper left', fontsize=9); ax.grid(alpha=0.3)

# Bottom left: residuals overlay
ax = axes[1, 0]
ax.axhline(0, color='k', lw=1, alpha=0.5)
ax.scatter(rt_b61, np.array(res_b61)*1000, c='tab:orange', s=50, marker='s',
           label=f'V6.1 corr - RT  (MAE {mae_b61_corr*1000:.1f}, mean {mean_b61_corr*1000:+.1f})', zorder=3)
ax.scatter(rt_b70, np.array(res_b70)*1000, c='tab:green', s=50, marker='o',
           label=f'V7.0 corr - RT  (MAE {mae_b70_corr*1000:.1f}, mean {mean_b70_corr*1000:+.1f})', zorder=3)
ax.set_xlabel('RealTech (cm$^{-1}$)'); ax.set_ylabel('DIY_corr - RT (mAU)')
ax.set_title('Residual comparison: V6.1 vs V7.0')
ax.legend(loc='best', fontsize=9); ax.grid(alpha=0.3)

# Bottom right: MAE bar chart
ax = axes[1, 1]
labels = ['V6.1\nraw', 'V6.1\ncorr', 'V7.0\nraw', 'V7.0\ncorr']
maes = [mae_b61_raw*1000, mae_b61_corr*1000, mae_b70_raw*1000, mae_b70_corr*1000]
colors = ['#8888cc', '#cc4444', '#8888cc', '#44aa44']
bars = ax.bar(labels, maes, color=colors)
for b, v in zip(bars, maes):
    ax.text(b.get_x()+b.get_width()/2, v+0.15, f'{v:.1f}', ha='center', fontsize=10)
ax.set_ylabel('Mean absolute error vs RealTech (mAU)')
ax.set_title('MAE: raw vs corrected, each firmware')
ax.grid(alpha=0.3, axis='y')

fig.suptitle('Head-to-head: Ver 6.1 (single-pt cal) vs Ver 7.0 (2-pt affine cal)  — Aug 12 1312',
             fontsize=13, y=1.00)
plt.tight_layout()
plt.savefig(HERE / 'plot_v61_vs_v70_head_to_head.png', dpi=140, bbox_inches='tight')
plt.close()

# --- Plot 3: Compare Ver 7.0 firmware behavior to hand-fit affine ---
# What if we FIT the best affine post-hoc to V7.0 raw vs RT? How does that
# compare to what the firmware ACTUALLY applied?
m_best, b_best, r2_best = fit_line(raw_b70, rt_b70)  # solve rt = m*raw + b
corr_optimal = np.array(raw_b70) * m_best + b_best
res_optimal = corr_optimal - np.array(rt_b70)
mae_optimal = float(np.mean(np.abs(res_optimal)))

print(f"\nBest possible affine (post-hoc least-squares on V7.0 raw vs RT):")
print(f"  m={m_best:.4f}, b={b_best:.5f} ({b_best*1000:+.2f} mAU), r2={r2_best:.4f}")
print(f"  best-affine MAE: {mae_optimal*1000:.1f} mAU  (vs firmware MAE {mae_b70_corr*1000:.1f})")

fig, ax = plt.subplots(figsize=(9, 6.5))
xmax = max(max(rt_b70), max(raw_b70), max(corr_b70)) * 1.1
xs = np.linspace(0, xmax, 200)
ax.plot(xs, xs, 'k--', lw=1, alpha=0.5, label='y = x (perfect)')
ax.scatter(rt_b70, raw_b70, c='tab:blue', s=55, label=f'DIY raw (slope {m_b70_raw:.3f})', zorder=3)
ax.scatter(rt_b70, corr_b70, c='tab:red', s=55, marker='s',
           label=f'DIY corr = m·raw+b  (firmware m={m_fw_b70:.3f}, b={b_fw_b70*1000:+.1f} mAU)  MAE={mae_b70_corr*1000:.1f}', zorder=3)
ax.scatter(rt_b70, corr_optimal, c='tab:green', s=55, marker='^',
           label=f'DIY corr (best-fit affine)  m={m_best:.3f}, b={b_best*1000:+.1f} mAU  MAE={mae_optimal*1000:.1f}', zorder=3)
ax.plot(xs, m_b70_raw*xs + b_b70_raw, 'tab:blue', alpha=0.35, lw=1)
ax.plot(xs, m_b70_corr*xs + b_b70_corr, 'tab:red', alpha=0.35, lw=1)
ax.set_xlabel('RealTech absorbance (cm$^{-1}$)')
ax.set_ylabel('DIY absorbance (cm$^{-1}$)')
ax.set_title('Ver 7.0 head-to-head data: firmware affine vs best-fit affine\n'
             '(shows how much of the error is calibration choice vs residual scatter)')
ax.set_xlim(0, xmax); ax.set_ylim(0, xmax); ax.set_aspect('equal')
ax.legend(loc='upper left', fontsize=9); ax.grid(alpha=0.3)
plt.tight_layout()
plt.savefig(HERE / 'plot_v70_firmware_vs_best_affine.png', dpi=140, bbox_inches='tight')
plt.close()

# --- Write summary.md and csv ---
summary_md = f"""# Aug 12 1312 — Ver 7.0 (2-pt affine) validation

Source: `data.xlsx`.

## Block A: Ver 7.0 on fluorescein solutions
(n = {len(rt_a)}, RT range {min(rt_a):.3f}–{max(rt_a):.3f} cm⁻¹)

The firmware was calibrated with 0.010 and 0.100 fluorescein anchors. Back-solved
from the (raw, corr) pairs: **m = {m_fw_a:.4f}, b = {b_fw_a*1000:+.2f} mAU**
(essentially pure multiplicative — b ≈ 0).

| metric | raw | corrected |
|---|---|---|
| slope vs RT | {m_a_raw:.3f} | {m_a_corr:.3f} |
| intercept | {b_a_raw*1000:+.2f} mAU | {b_a_corr*1000:+.2f} mAU |
| R² | {r2_a_raw:.4f} | {r2_a_corr:.4f} |
| MAE vs RT | {mae_a_raw*1000:.2f} mAU | **{mae_a_corr*1000:.2f} mAU** |
| max err | {max_a_raw*1000:.2f} mAU | {max_a_corr*1000:.2f} mAU |
| mean signed | {mean_a_raw*1000:+.2f} mAU | {mean_a_corr*1000:+.2f} mAU |
| all-positive resid? | {allpos_a_raw} | **{allpos_a_corr}** |

Compare to Aug 10 1533 fluor-only (Ver 6.1, 1-pt cal): MAE = 4.9 mAU, all-positive.
Here MAE = **{mae_a_corr*1000:.1f} mAU** and residuals swing both signs.

## Block B: Ver 6.1 vs Ver 7.0 head-to-head
(separate sample lists for each firmware; both against RealTech)

### Ver 6.1 (n = {len(rt_b61)}, 1-pt cal, k ≈ {k_b61_avg:.3f})

| metric | raw | corrected |
|---|---|---|
| slope vs RT | {m_b61_raw:.3f} | {m_b61_corr:.3f} |
| MAE vs RT | {mae_b61_raw*1000:.2f} mAU | **{mae_b61_corr*1000:.2f} mAU** |
| max err | {max_b61_raw*1000:.2f} mAU | {max_b61_corr*1000:.2f} mAU |
| mean signed | {mean_b61_raw*1000:+.2f} mAU | {mean_b61_corr*1000:+.2f} mAU |
| all-positive resid? | — | {allpos_b61} |

### Ver 7.0 (n = {len(rt_b70)}, 2-pt affine)

Firmware applied: **m = {m_fw_b70:.4f}, b = {b_fw_b70*1000:+.2f} mAU** (recovered
from the (raw, corr) pairs).

| metric | raw | corrected |
|---|---|---|
| slope vs RT | {m_b70_raw:.3f} | {m_b70_corr:.3f} |
| MAE vs RT | {mae_b70_raw*1000:.2f} mAU | **{mae_b70_corr*1000:.2f} mAU** |
| max err | {max_b70_raw*1000:.2f} mAU | {max_b70_corr*1000:.2f} mAU |
| mean signed | {mean_b70_raw*1000:+.2f} mAU | {mean_b70_corr*1000:+.2f} mAU |
| all-positive resid? | — | {allpos_b70} |

### Best possible affine (post-hoc least-squares) on same V7.0 raw values
m = {m_best:.4f}, b = {b_best*1000:+.2f} mAU → MAE = **{mae_optimal*1000:.2f} mAU**.

The gap (firmware {mae_b70_corr*1000:.1f} vs best possible {mae_optimal*1000:.1f}) is
the calibration-choice cost — how much the on-device 2-pt anchors differ from the
optimal fit over the whole sample range. The remainder is irreducible scatter.

## Plots
- `plot_v70_fluor_only.png` — Block A scatter + residuals
- `plot_v61_vs_v70_head_to_head.png` — Block B 4-panel comparison
- `plot_v70_firmware_vs_best_affine.png` — how close firmware got to the optimal 2-pt affine
"""

(HERE / "summary.md").write_text(summary_md)

with (HERE / "summary.csv").open("w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["block", "firmware", "n", "slope_raw", "slope_corr",
                "mae_raw_mAU", "mae_corr_mAU", "max_corr_mAU",
                "mean_signed_corr_mAU", "all_positive_corr",
                "firmware_m", "firmware_b_mAU"])
    w.writerow(["A_fluor_only", "v7.0", len(rt_a), f"{m_a_raw:.3f}", f"{m_a_corr:.3f}",
                f"{mae_a_raw*1000:.2f}", f"{mae_a_corr*1000:.2f}", f"{max_a_corr*1000:.2f}",
                f"{mean_a_corr*1000:+.2f}", allpos_a_corr,
                f"{m_fw_a:.4f}", f"{b_fw_a*1000:+.2f}"])
    w.writerow(["B_head_to_head", "v6.1", len(rt_b61), f"{m_b61_raw:.3f}", f"{m_b61_corr:.3f}",
                f"{mae_b61_raw*1000:.2f}", f"{mae_b61_corr*1000:.2f}", f"{max_b61_corr*1000:.2f}",
                f"{mean_b61_corr*1000:+.2f}", allpos_b61, "", ""])
    w.writerow(["B_head_to_head", "v7.0", len(rt_b70), f"{m_b70_raw:.3f}", f"{m_b70_corr:.3f}",
                f"{mae_b70_raw*1000:.2f}", f"{mae_b70_corr*1000:.2f}", f"{max_b70_corr*1000:.2f}",
                f"{mean_b70_corr*1000:+.2f}", allpos_b70,
                f"{m_fw_b70:.4f}", f"{b_fw_b70*1000:+.2f}"])

print("\nWrote:")
for p in sorted(HERE.glob("*")):
    if p.name != "analyze.py":
        print(f"  {p.name}")
